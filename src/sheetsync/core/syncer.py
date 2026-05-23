from __future__ import annotations

import csv
import hashlib
import json
import socket
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable

import gspread
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from .google_auth import AuthError, get_client
from .storage import AppConfig, SyncPairConfig, clear_lock, create_lock, ensure_pair, load_queue, save_config, save_queue


StatusCallback = Callable[[dict], None]


class SyncError(RuntimeError):
    pass


@dataclass
class TableData:
    headers: list[str]
    rows: list[dict[str, object]]


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def internet_available(host: str = "www.googleapis.com", port: int = 443, timeout: float = 3) -> bool:
    try:
        with socket.create_connection((host, port), timeout=timeout):
            return True
    except OSError:
        return False


def _normalize(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_excel(path: str) -> TableData:
    file_path = Path(path)
    if not file_path.exists():
        raise FileNotFoundError(path)
    last_error: PermissionError | None = None
    for _ in range(5):
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            values = list(sheet.iter_rows(values_only=True))
            workbook.close()
            if not values:
                return TableData([], [])
            headers = [_normalize(v) for v in values[0]]
            rows = []
            for raw in values[1:]:
                row = {headers[i]: raw[i] if i < len(raw) else "" for i in range(len(headers)) if headers[i]}
                if any(_normalize(v) for v in row.values()):
                    rows.append(row)
            return TableData(headers, rows)
        except PermissionError as exc:
            last_error = exc
            time.sleep(2)
    raise SyncError("Waiting for Excel to release file") from last_error


def write_excel(path: str, table: TableData) -> None:
    last_error: PermissionError | None = None
    for _ in range(5):
        workbook = Workbook()
        try:
            sheet = workbook.active
            sheet.append(table.headers)
            for row in table.rows:
                sheet.append([row.get(header, "") for header in table.headers])
            workbook.save(path)
            return
        except PermissionError as exc:
            last_error = exc
            time.sleep(2)
        finally:
            workbook.close()
    raise SyncError("Waiting for Excel to release file") from last_error


def read_sheet(client: gspread.Client, config: AppConfig | SyncPairConfig) -> tuple[gspread.Worksheet, TableData]:
    spreadsheet = client.open_by_key(config.sheet_id)
    worksheet = spreadsheet.worksheet(config.worksheet_name) if config.worksheet_name else spreadsheet.sheet1
    values = worksheet.get_all_values()
    if not values:
        return worksheet, TableData([], [])
    headers = [_normalize(v) for v in values[0]]
    rows = []
    for raw in values[1:]:
        row = {headers[i]: raw[i] if i < len(raw) else "" for i in range(len(headers)) if headers[i]}
        if any(_normalize(v) for v in row.values()):
            rows.append(row)
    return worksheet, TableData(headers, rows)


def write_sheet(worksheet: gspread.Worksheet, table: TableData) -> None:
    values = [table.headers] + [[row.get(header, "") for header in table.headers] for row in table.rows]
    worksheet.clear()
    if values:
        worksheet.update(values, "A1")


def rows_by_key(table: TableData, key: str, side: str) -> dict[str, dict[str, object]]:
    if key not in table.headers:
        raise SyncError(f"Match column '{key}' is missing in {side}.")
    mapped = {}
    for row in table.rows:
        value = _normalize(row.get(key))
        if value:
            mapped[value] = row
    return mapped


def rows_by_position(table: TableData) -> dict[str, dict[str, object]]:
    return {str(index): row for index, row in enumerate(table.rows, start=1)}


def merge_headers(a: list[str], b: list[str]) -> list[str]:
    merged = []
    for header in a + b:
        if header and header not in merged:
            merged.append(header)
    return merged


def row_map(table: TableData, key: str, config: AppConfig | SyncPairConfig, side: str) -> dict[str, dict[str, object]]:
    if config.match_mode == "Column value":
        return rows_by_key(table, key, side)
    return rows_by_position(table)


def ordered_keys(excel_map: dict[str, dict[str, object]], sheet_map: dict[str, dict[str, object]], config: AppConfig | SyncPairConfig) -> list[str]:
    keys = set(excel_map) | set(sheet_map)
    if config.match_mode == "Column value":
        return sorted(keys)
    return sorted(keys, key=lambda value: int(value))


def compute_sheet_hash(rows: list[dict]) -> str:
    serialized = json.dumps([sorted(row.items()) for row in rows], sort_keys=True)
    return hashlib.sha1(serialized.encode()).hexdigest()


def apply_column_mappings(table: TableData, mappings: dict[str, str]) -> TableData:
    """Rename Excel-side columns to their Sheets-side names. No-op for empty mappings."""
    if not mappings:
        return table
    new_headers = [mappings.get(header, header) for header in table.headers]
    new_rows = [{mappings.get(key, key): value for key, value in row.items()} for row in table.rows]
    return TableData(new_headers, new_rows)


def reverse_column_mappings(table: TableData, mappings: dict[str, str]) -> TableData:
    """Inverse of apply_column_mappings — used to translate back before writing Excel."""
    if not mappings:
        return table
    return apply_column_mappings(table, {sheets: excel for excel, sheets in mappings.items()})


def compute_merge(excel: TableData, sheet: TableData, key: str, config: AppConfig | SyncPairConfig) -> tuple[TableData, TableData, dict[str, int], list[dict]]:
    excel_map = row_map(excel, key, config, "Excel file")
    sheet_map = row_map(sheet, key, config, "Google Sheet")
    headers = merge_headers(excel.headers, sheet.headers)
    all_keys = ordered_keys(excel_map, sheet_map, config)
    stats = {"added": 0, "modified": 0, "deleted": 0, "conflicts": 0, "rows": 0}
    conflict_log: list[dict] = []
    merged_rows: list[dict[str, object]] = []
    direction = config.sync_direction
    for item_key in all_keys:
        erow = excel_map.get(item_key)
        srow = sheet_map.get(item_key)
        if erow and not srow:
            # Excel has a row Sheets doesn't — Sheets->Excel will erase it, else it propagates as an add.
            chosen = erow
            bucket = "deleted" if direction == "Sheets -> Excel" else "added"
            stats[bucket] += 1
        elif srow and not erow:
            chosen = srow
            bucket = "deleted" if direction == "Excel -> Sheets" else "added"
            stats[bucket] += 1
        elif erow == srow:
            chosen = erow or srow or {}
        else:
            stats["modified"] += 1
            stats["conflicts"] += 1
            if config.conflict_resolution == "Sheets wins":
                chosen = srow or {}
                resolved_to = "sheets"
            else:
                chosen = erow or {}
                resolved_to = "excel"
            for col in headers:
                excel_val = _normalize((erow or {}).get(col))
                sheet_val = _normalize((srow or {}).get(col))
                if excel_val != sheet_val:
                    conflict_log.append({
                        "key": item_key,
                        "col": col,
                        "excel_val": excel_val,
                        "sheet_val": sheet_val,
                        "resolved_to": resolved_to,
                    })
        merged_rows.append({header: chosen.get(header, "") for header in headers})
    stats["rows"] = stats["added"] + stats["modified"] + stats["deleted"]
    merged = TableData(headers, merged_rows)
    if config.sync_direction == "Excel -> Sheets":
        return excel, excel, stats, conflict_log
    if config.sync_direction == "Sheets -> Excel":
        return sheet, sheet, stats, conflict_log
    return merged, merged, stats, conflict_log


class SyncEngine:
    def __init__(self, config: AppConfig, callback: StatusCallback | None = None):
        self.config = config
        self.callback = callback or (lambda event: None)
        self.conflict_logs: dict[str, list[dict]] = {}

    def emit(self, kind: str, message: str, **data: object) -> None:
        self.callback({"kind": kind, "message": message, **data})

    def sync_now(self, reason: str = "manual", pair_id: str | None = None) -> dict[str, int]:
        pair = ensure_pair(self.config, pair_id)
        if self.config.paused or pair.paused:
            raise SyncError("Sync is paused.")
        if not pair.excel_path:
            raise SyncError("Choose an Excel file before syncing.")
        if not pair.sheet_id:
            raise SyncError("Connect a Google Sheet before syncing.")
        if not internet_available():
            queue = load_queue()
            queue.append({"created_at": now_iso(), "reason": reason, "pair_id": pair.id, "excel_path": pair.excel_path})
            save_queue(queue)
            pair.queued_changes = sum(1 for item in queue if item.get("pair_id") == pair.id)
            self.config.queued_changes = len(queue)
            save_config(self.config)
            raise ConnectionError("Offline. Changes queued and will sync when the connection returns.")
        create_lock()
        self.emit("status", "Syncing...", state="syncing")
        try:
            excel = read_excel(pair.excel_path)
            client = get_client(interactive=False)
            worksheet, sheet = read_sheet(client, pair)
            mappings = pair.column_mappings or {}
            excel_mapped = apply_column_mappings(excel, mappings)
            match_key = mappings.get(pair.match_column, pair.match_column)
            excel_out_mapped, sheet_out, stats, conflict_log = compute_merge(excel_mapped, sheet, match_key, pair)
            excel_out = reverse_column_mappings(excel_out_mapped, mappings)
            if mappings:
                # Conflict log columns are in Sheets-space — translate back so the UI shows Excel column names.
                reversed_map = {sheets: excel for excel, sheets in mappings.items()}
                conflict_log = [{**entry, "col": reversed_map.get(entry["col"], entry["col"])} for entry in conflict_log]
            should_write_excel = pair.sync_direction in ("Bidirectional", "Sheets -> Excel")
            if reason == "Excel" and pair.sync_direction == "Bidirectional":
                should_write_excel = False
            if should_write_excel:
                write_excel(pair.excel_path, excel_out)
            if pair.sync_direction in ("Bidirectional", "Excel -> Sheets"):
                write_sheet(worksheet, sheet_out)
            pair.last_sync_iso = now_iso()
            if reason == "Excel":
                pair.last_edited_side = "excel"
            elif reason == "sheets_poll":
                pair.last_edited_side = "sheets"
            pair.queued_changes = 0
            pair.stats["col_count"] = len(excel.headers)
            pair.stats["total_syncs"] += 1
            pair.stats["rows_synced"] += stats["rows"]
            pair.stats["conflicts_resolved"] += stats["conflicts"]
            self.conflict_logs[pair.id] = (self.conflict_logs.get(pair.id, []) + conflict_log)[-100:]
            self.config.last_sync_iso = pair.last_sync_iso
            self.config.queued_changes = sum(item.queued_changes for item in self.config.pairs)
            self.config.stats["total_syncs"] += 1
            self.config.stats["rows_synced"] += stats["rows"]
            self.config.stats["conflicts_resolved"] += stats["conflicts"]
            remaining_queue = [item for item in load_queue() if item.get("pair_id") != pair.id]
            save_queue(remaining_queue)
            save_config(self.config)
            self.emit("synced", f"Synced {stats['rows']} rows", stats=stats)
            return stats
        except FileNotFoundError as exc:
            pair.stats["errors"] += 1
            self.config.stats["errors"] += 1
            save_config(self.config)
            raise SyncError("Excel file not found. Locate the file to resume syncing.") from exc
        except PermissionError as exc:
            pair.stats["errors"] += 1
            self.config.stats["errors"] += 1
            save_config(self.config)
            raise SyncError("Waiting for Excel to release file") from exc
        except AuthError:
            pair.stats["errors"] += 1
            self.config.stats["errors"] += 1
            save_config(self.config)
            raise
        except gspread.SpreadsheetNotFound as exc:
            pair.stats["errors"] += 1
            self.config.stats["errors"] += 1
            save_config(self.config)
            raise SyncError("Google Sheet was deleted or access was removed. Update the Sheet URL.") from exc
        finally:
            clear_lock()

    def poll_sheet_hash(self, pair_id: str) -> str:
        pair = ensure_pair(self.config, pair_id)
        try:
            client = get_client(interactive=False)
            _, sheet = read_sheet(client, pair)
            return compute_sheet_hash(sheet.rows)
        except Exception:
            return ""


def export_activity_csv(path: str, entries: list[dict]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["time", "kind", "message", "rows"])
        writer.writeheader()
        for entry in entries:
            writer.writerow(
                {
                    "time": entry.get("time", ""),
                    "kind": entry.get("kind", ""),
                    "message": entry.get("message", ""),
                    "rows": entry.get("rows", ""),
                }
            )
